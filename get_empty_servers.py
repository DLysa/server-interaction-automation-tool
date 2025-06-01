import requests
import pyperclip
import keyboard
import time
import os
import json
import mouse
import pytesseract
import re
from datetime import datetime, timedelta
from PIL import Image, ImageGrab, ImageEnhance, ImageFilter
import threading
import cv2
import numpy as np
import os
import json
from datetime import datetime
from openpyxl import Workbook, load_workbook

# Settings
url = "https://xplay.gg/api/play/getAllServers"
output_file = "empty_servers.txt"
used_file = "used_servers.json"
log_file = "server_log6txt"
time_limit = timedelta(minutes=51)
tesseract_cmd = r"C:\\Program Files\\Tesseract-OCR\\tesseract.exe"
pytesseract.pytesseract.tesseract_cmd = tesseract_cmd
stop_macro = False
warmup_found = False
autoskip_count = 0
warmup_found_count = 0
recent_autoskips = set()

PERSISTENT_DIR = "persistent"
SESSION_DIR = "session"
os.makedirs(PERSISTENT_DIR, exist_ok=True)
os.makedirs(SESSION_DIR, exist_ok=True)

PERSISTENT_FILE = os.path.join(PERSISTENT_DIR, "autoskips.xlsx")
SESSION_FILE = os.path.join(SESSION_DIR, f"autoskips_session_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx")


def preprocess_for_ocr(pil_image):
    open_cv_image = np.array(pil_image)

    # Only convert if image has 3 channels
    if len(open_cv_image.shape) == 3 and open_cv_image.shape[2] == 3:
        open_cv_image = cv2.cvtColor(open_cv_image, cv2.COLOR_RGB2GRAY)

    open_cv_image = cv2.adaptiveThreshold(
        open_cv_image, 255, cv2.ADAPTIVE_THRESH_GAUSSIAN_C,
        cv2.THRESH_BINARY, 11, 2
    )

    return Image.fromarray(open_cv_image)


def init_excel_file(path):
    if not os.path.exists(path):
        wb = Workbook()
        ws = wb.active
        ws.append(["IP:Port", "Skip Count", "Skip Times"])
        wb.save(path)


init_excel_file(PERSISTENT_FILE)
init_excel_file(SESSION_FILE)


# Update skip data
def log_autoskip(ip_port):
    now_str = datetime.now().isoformat()
    for file_path in [PERSISTENT_FILE, SESSION_FILE]:
        wb = load_workbook(file_path)
        ws = wb.active

        # Find the row with the IP:Port
        for row in ws.iter_rows(min_row=2, values_only=False):
            cell_ip = row[0]
            cell_count = row[1]
            cell_times = row[2]
            if cell_ip.value == ip_port:
                cell_count.value += 1
                cell_times.value += f", {now_str}"
                break
        else:
            # Not found, add new row
            ws.append([ip_port, 1, now_str])

        wb.save(file_path)


# Your original log_event function (with path to text log file)
log_file = "server_log.txt"


def log_event(ip_port, action, reason=None, unban_time=None, join_time=None, warmup_result=None):
    log_data = {
        "timestamp": datetime.now().isoformat(),
        "ip_port": ip_port,
        "action": action,
        "reason": reason,
        "unban_time": unban_time.isoformat() if unban_time else None,
        "last_join_attempt": join_time.isoformat() if join_time else None,
        "warmup_result": warmup_result
    }
    with open(log_file, "a") as f:
        f.write(json.dumps(log_data) + "\n")


def load_used_servers():
    if not os.path.exists(used_file):
        return {}
    with open(used_file, "r") as f:
        try:
            data = json.load(f)
        except json.JSONDecodeError:
            return {}
    return data


def save_used_servers(used):
    with open(used_file, "w") as f:
        json.dump(used, f)


def cleanup_used_servers(used):
    now = datetime.now()
    still_used = {}
    expired = []

    for ip, timestamp in used.items():
        try:
            time_diff = now - datetime.fromisoformat(timestamp)
            if time_diff < time_limit:
                still_used[ip] = timestamp
            else:
                expired.append(ip)
        except Exception as e:
            print(f"[ERROR] Parsing timestamp for {ip}: {e}")
            expired.append(ip)

    # Save updated used list
    save_used_servers(still_used)

    # 🔧 Remove expired from empty list if present
    if os.path.exists(output_file):
        with open(output_file, "r") as f:
            empty = f.read().splitlines()
        new_empty = [line for line in empty if all(ip not in line for ip in expired)]
        if len(new_empty) != len(empty):
            with open(output_file, "w") as f:
                f.write("\n".join(new_empty) + "\n")
            print(f"[CLEANUP] Removed expired servers from empty list")

    return still_used


def fetch_and_filter_servers(used_servers):
    try:
        print("[INFO] Fetching server data...")
        response = requests.get(url)
        response.raise_for_status()

        data = response.json()
        servers = data.get("serversList", [])

        print("[INFO] Filtering servers...")
        filtered = []
        seen = set()
        now = datetime.now()

        for server in servers:
            ip = server.get("IP")
            port = server.get("Port")
            if not ip or not port:
                continue

            ip_port = f"{ip}:{port}"
            is_populated = server.get("Online", 1) != 0 or server.get("GameModeID") not in [27, 28]

            if is_populated:
                used_servers[ip_port] = now.isoformat()
                log_event(ip_port, "banned", reason="populated", unban_time=now + time_limit)

                if os.path.exists(output_file):
                    with open(output_file, "r") as f:
                        empty = f.read().splitlines()
                    new_empty = [line for line in empty if ip_port not in line]
                    if len(new_empty) != len(empty):
                        with open(output_file, "w") as f:
                            f.write("\n".join(new_empty) + "\n")
                        print(f"[INFO] Removed banned server from empty list.")
                continue

            if ip_port not in seen and ip_port not in used_servers:
                seen.add(ip_port)
                filtered.append(f"connect {ip_port}")

        save_used_servers(used_servers)

        if filtered:
            if os.path.exists(output_file):
                with open(output_file, "r") as f:
                    existing_servers = f.read().splitlines()
            else:
                existing_servers = []

            combined_servers = existing_servers + filtered
            combined_servers = list(dict.fromkeys(combined_servers))

            with open(output_file, "w") as f:
                f.write("\n".join(combined_servers) + "\n")

            print(f"[SUCCESS] {len(filtered)} new empty servers appended. Total now: {len(combined_servers)}")
        else:
            print("[WARNING] No matching empty servers found.")

        recent_autoskips.clear()

    except Exception as e:
        print(f"[ERROR] {e}")


def capture_center_bottom_screen():
    screen = ImageGrab.grab()
    width, height = screen.size

    left = width * 0.46
    top = height * 0.71
    right = width * 0.54
    bottom = height * 0.74

    cropped = screen.crop((left, top, right, bottom))

    # Increase size for better OCR accuracy
    new_size = (cropped.width * 4, cropped.height * 4)

    try:
        resample_filter = Image.Resampling.LANCZOS  # Pillow >= 10
    except AttributeError:
        resample_filter = Image.LANCZOS if hasattr(Image, 'LANCZOS') else 1

    cropped = cropped.resize(new_size, resample=resample_filter)

    # Enhance contrast and brightness
    enhancer = ImageEnhance.Contrast(cropped)
    cropped = enhancer.enhance(2.5)

    enhancer = ImageEnhance.Brightness(cropped)
    cropped = enhancer.enhance(1.2)

    # Convert to grayscale and apply binary threshold/4/4/4/\
    cv_img = np.array(cropped.convert("L"))  # grayscale
    _, thresh = cv2.threshold(cv_img, 180, 255, cv2.THRESH_BINARY_INV)
    cropped = Image.fromarray(thresh)

    return cropped


def check_warmup_text(timeout=6):
    global warmup_found
    time.sleep(1)
    start_time = time.time()
    while time.time() - start_time < timeout:
        cropped = capture_center_bottom_screen()
        config = r'--oem 3 --psm 7'
        raw_text = pytesseract.image_to_string(cropped, config=config)

        print(f"[OCR] Raw detected text: {raw_text.strip()}")

        # Basic cleaning
        text = raw_text
        text = text.replace("|", "")
        text = text.replace(".", ":")  # Sometimes colon is detected as dot
        text = re.sub(r'\s+', ' ', text)

        print(f"[OCR] Cleaned text: {text.strip()}")

        # Flexible match: "Warmup 0:25" or similar
        if (re.search(r"(?i)warm.?up\s*0[:.]?\d{1,2}", text) or
                re.search(r'(?i)warmup\s*\d{1,2}', text) or
                re.search(r'(?i)warm[\s_]*up[\s_]*[0o]?[:\s]?\d{1,3}', text)
        ):
            print("[SUCCESS] Warmup found!")
            warmup_found = True
            return True

        time.sleep(0.1)

    print("[INFO] Warmup not found after timeout.")
    return False


def run_macro():
    global stop_macro, warmup_found, step

    sequence = [
        ("b", "5", "2"), ("b", "b"), (0.01,), ("4",), (1.2,), ("/",),
        ("b", "5", "3"), ("b", "b"), (0.01,), ("4",), (1.15,), ("/",),
        ("b", "5", "4"), ("b", "b"), (0.01,), ("4",), (1.1,), ("/",),
        "hold_s",
        ("b", "5", "5"), ("b", "b"), (0.01,), ("4",), (1.1,), ("/",),
        ("b", "5", "1"), ("b", "b"), (0.01,), ("4",), (1.06,), ("/",),
        ("b", "5", "1"), ("b", "b"), (0.01,), ("4",), (1.06,), ("/",),
        ("b", "5", "1"), ("b", "b"), (0.01,), ("4",), (1.06,), ("/",),
        ("b", "5", "1"), ("b", "b"), (0.01,), ("4",), (1.06,), ("/",),
        ("b", "5", "1"), ("b", "b"), (0.01,), ("4",), (1.06,), ("/",),
        ("b", "5", "1"), ("b", "b"), (0.01,), ("4",), (1.1,), ("/",),
        ("b", "5", "2"), ("b", "b"), (0.1,), ("4",), (1.2,), ("/",),
        ("b", "5", "3"), ("b", "b"), (0.01,), ("4",), (1.1,), ("/",),
        ("b", "5", "4"), ("b", "b"), (0.01,), ("4",), (1.1,), ("/",),
        ("b", "5", "5"), ("b", "b"), (0.01,), ("4",), (1.1,), ("/",),
        "hold_w",
        ("b", "5", "2"), ("b", "b"), (0.1,), ("4",), (1.2,), ("/",),
        ("b", "5", "5"), ("b", "b"), (0.01,), ("4",), (1.1,), ("/",),
        ("b", "5", "3"), ("b", "b"), (0.01,), ("4",), (1.1,), ("/",),
        ("b", "5", "5"), ("b", "b"), (0.01,), ("4",), (1.1,), ("/",),
        ("b", "5", "2"), ("b", "b"), (0.1,), ("4",), (1.2,), ("/",),
        ("b", "5", "5"), ("b", "b"), (0.01,), ("4",), (1.1,), ("/",),
        ("b", "5", "3"), ("b", "b"), (0.01,), ("4",), (1.1,), ("/",),
        ("b", "5", "4"), ("b", "b"), (0.01,), ("4",), (1.1,), ("/",),
        ("b", "5", "2", "3", "4"), ("b", "b"), (0.01,), ("4",), (1.1,), ("/",),
        ("4",), (1.1,), ("/",),
        ("4",), (1.1,), ("/",),
        ("4",), (1.1,), ("/",)
    ]

    for step in sequence:
        if stop_macro:
            print("[INFO] Macro forcibly stopped because no Warmup detected.")
            break

        if step == "hold_s":
            keyboard.press('s')
            time.sleep(0.1)
            keyboard.release('s')
            continue

        if step == "hold_w":
            keyboard.press('w')
            time.sleep(0.1)
            keyboard.release('w')
            continue

        for action in step:
            if isinstance(action, (int, float)):
                time.sleep(action)
            else:
                keyboard.press_and_release(action)

        time.sleep(0.01)


print("[INFO] Macro finished normally.")


def check_warmup_text_monitor():
    global warmup_found
    start_time = time.time()

    while time.time() - start_time < 6:  # 12 seconds timeout for finding warmup
        cropped = capture_center_bottom_screen()
        config = r'--oem 3 --psm 6'
        raw_text = pytesseract.image_to_string(cropped, config=config)

        text = raw_text
        text = text.replace("|", "").replace(".", "").replace(")", "").replace("(", "")
        text = re.sub(r'\s+', ' ', text)

        # if re.search(r"Warmup(\s+Ending)?\s*\d{1,2}:\d{2}", text, re.IGNORECASE):
        if re.search(r"Warm[uvnp]+(\s+Ending)?\s*\d{1,2}[:\.]\d{2}", text, re.IGNORECASE):
            print("[SUCCESS] Warmup detected during macro!")
            warmup_found = True
            return

        time.sleep(0.1)

    print("[WARNING] Warmup NOT detected after timeout.")
    warmup_found = False


def start_macro_and_ocr(ip_port):
    global stop_macro, warmup_found, autoskip_count, warmup_found_count
    stop_macro = False
    warmup_found = False

    macro_thread = threading.Thread(target=run_macro)
    ocr_thread = threading.Thread(target=check_warmup_text)

    macro_thread.start()
    ocr_thread.start()

    ocr_thread.join()

    if not warmup_found:
        stop_macro = True
        autoskip_count += 1
        recent_autoskips.add(ip_port)
        screenshot = capture_center_bottom_screen()
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"screenshots/autoskip_{timestamp}.png"
        os.makedirs("screenshots", exist_ok=True)
        screenshot.save(filename)
        print(f"[SCREENSHOT] Saved screenshot before autoskip to {filename}")

        print(f"[INFO] Forcing macro to stop because warmup not found. [AutoSkips: {autoskip_count}]")
        log_autoskip(ip_port)
    else:
        warmup_found_count += 1
        print(f"[INFO] Warmup successfully detected. [Warmups Found: {warmup_found_count}]")

    macro_thread.join()

    print("[INFO] Both Macro and OCR threads finished.")


def interact_with_servers():
    last_activity = time.time()
    copy_count = 4
    forced_skip = False

    used_servers = load_used_servers()

    while True:
        used_servers = cleanup_used_servers(used_servers)

        if not os.path.exists(output_file) or os.stat(output_file).st_size == 0:
            print("[INFO] Server list is empty or missing. Fetching new server list...")
            fetch_and_filter_servers(used_servers)

        with open(output_file, "r") as f:
            servers = f.read().splitlines()

        if not servers:
            print("[INFO] No servers left. Retrying fetch...")
            time.sleep(5)
            continue

        if not forced_skip:
            print("[INFO] Waiting for '[' key to proceed... (auto after 3 min idle)")
            while True:
                event = keyboard.read_event(suppress=False)
                if event.name == "[" and event.event_type == "down":
                    break
                if time.time() - last_activity > 180:
                    print("[INFO] Idle timeout reached. Auto-triggering '[' key.")
                    keyboard.press_and_release("[")
                    last_activity = time.time()
                    break
                if event.event_type == "down":
                    last_activity = time.time()
        else:
            forced_skip = False

        command = servers.pop(0)
        ip_port = command.replace("connect ", "").strip()

        if ip_port in recent_autoskips:
            print(f"[SKIP] Skipping recently autoskipped server: {ip_port}")
            continue

        used_servers[ip_port] = datetime.now().isoformat()
        save_used_servers(used_servers)

        join_time = datetime.now()
        log_event(ip_port, "joined", reason="joined", unban_time=join_time + time_limit, join_time=join_time)

        with open(output_file, "w") as f:
            f.write("\n".join(servers))

        pyperclip.copy(command)
        keyboard.press_and_release("\\")
        time.sleep(0.1)

        copy_count += 1

        if copy_count % 5 == 0:
            fetch_and_filter_servers(used_servers)
            with open(output_file, "r") as f:
                servers = f.read().splitlines()
            time.sleep(0.1)
            keyboard.press_and_release("ctrl+a")
            keyboard.press_and_release("backspace")
            keyboard.write("bind / +attack")
            keyboard.press_and_release("enter")
            print("[INFO] Server list refreshed and attack button was bound")

        time.sleep(0.2)
        keyboard.press_and_release("ctrl+a")
        keyboard.press_and_release("backspace")
        keyboard.press_and_release("ctrl+v")
        keyboard.press_and_release("enter")
        keyboard.press_and_release("esc")

        print(f"[SEND] Sent command: {command}")

        print("[INFO] Waiting 8s for server to load...")
        time.sleep(8.3)

        keyboard.press_and_release("m")
        time.sleep(0.25)
        mouse.click('left')
        time.sleep(0.7)
        print("Macro started")

        start_macro_and_ocr(ip_port)
        success_rate = warmup_found_count / (warmup_found_count + autoskip_count) * 100
        print(
            f"[STATUS] Total Autoskips: {autoskip_count}, Total Warmups Found: {warmup_found_count} ------- {success_rate}% ")

        warmup_result = "warmup_found" if warmup_found else "no_warmup"
        log_event(ip_port, "joined_result", reason=warmup_result, join_time=join_time, warmup_result=warmup_result)

        if not warmup_found:
            print("[WARNING] Warmup not found. Skipping server.")
            forced_skip = True
            continue

        time.sleep(1.5)
        forced_skip = True


def start_periodic_refetch(used_servers, interval_seconds=150):
    def refetch_loop():
        while True:
            print("[TIMER] Periodic refetching of servers...")
            fetch_and_filter_servers(used_servers)
            time.sleep(interval_seconds)

    refetch_thread = threading.Thread(target=refetch_loop, daemon=True)
    refetch_thread.start()


if __name__ == "__main__":
    choice = input(f"Do you want to fetch a NEW server list and overwrite '{output_file}'? (y/n): ").strip().lower()
    used_servers = load_used_servers()
    if choice == "y":
        fetch_and_filter_servers(used_servers)

    # 🆕 Start periodic refetch every 3 minutes
    start_periodic_refetch(used_servers)

    interact_with_servers()
